import os
from openpyxl import Workbook, load_workbook
from django.shortcuts import render, redirect
from app.models import activity, Review
from .whatsapp import send_whatsapp_message  # import the helper function

# Login view
def login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        # Save login data
        login_entry = activity(username=username, password=password)
        login_entry.save()

        return redirect("index")
    return render(request, "login.html")


# Index view
def index(request):
    return render(request, "index.html")


# Review form view
def review(request):
    if request.method == "POST":
        name = request.POST.get("name")
        email = request.POST.get("email")
        phone = request.POST.get("phone")
        drc = request.POST.get("drc")

        review_entry = Review(name=name, email=email, phone=phone, drc=drc)
        review_entry.save()

        # WhatsApp message
        seller_phone = "+91XXXXXXXXX"  # Replace with actual seller WhatsApp number
        message = (
            f"📝 New Review Received:\n"
            f"👤 Name: {name}\n"
            f"📧 Email: {email}\n"
            f"📞 Phone: {phone}\n"
            f"💬 Review: {drc}"
        )

        # Send the message
        send_whatsapp_message(seller_phone, message)
        file_path = "reviews.xlsx"
        file_exists = os.path.exists(file_path)

        if not file_exists:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Name", "Email", "Phone", "Review"])
        else:
            workbook = load_workbook(file_path)
            sheet = workbook.active

        sheet.append([name, email, phone, drc])
        workbook.save(file_path)
        return redirect("index")

    return render(request, "review.html")
